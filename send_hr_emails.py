"""
HR Mail Merge — personalized via Gemini (rotating across models + API keys), sent via SMTP.

WHAT THIS DOES
1. Reads the HR contact database (xlsx, sheet "HR Contacts": S.No/Name/Title/Company/Category/Email)
2. For each contact, asks Gemini for a short, genuine-sounding opening line referencing their
   Title/Company — then drops it into a fixed template. Requests round-robin across
   GEMINI_MODELS and GOOGLE_API_KEYS so no single model/key gets hammered or rate-limited.
3. Sends via SMTP with your resume attached
4. Logs every send to sent_log.json (with timestamp/name/company) so re-running the script
   never double-emails anyone, and writes stats.json for the status dashboard
5. Throttles sends (daily cap + delay) so you don't get flagged as spam

BEFORE YOU RUN THIS
1. `pip install openpyxl requests`
2. Fill in the CONFIG block below — especially SMTP_USER, SMTP_PASS, RESUME_PATH, EXCEL_PATH
3. Gmail: you need an "App Password", not your normal password.
   Google Account -> Security -> 2-Step Verification -> App Passwords -> generate one for "Mail"
4. Set GOOGLE_API_KEY_1..16 (or GOOGLE_API_KEY/GEMINI_API_KEY) so there's at least one Gemini key
5. Run with DRY_RUN = True first (or env DRY_RUN=true). It will print 3 sample personalized
   emails and NOT send anything.
6. When happy, set DRY_RUN = False and rerun. It's safe to Ctrl+C and rerun anytime —
   sent_log.json means nobody gets double-emailed.

REALISTIC PACING
~1,800 contacts at a safe daily cap of ~40/day is ~45 days. Don't be tempted to blast
1000+/day from a personal Gmail — you'll get rate-limited or your account flagged, and
NO email will go out until Google unblocks you. Slow and steady actually gets read.
"""

import json
import os
import re
import smtplib
import ssl
import sys
import time
from datetime import datetime, timezone
from typing import Optional
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path

import requests
from openpyxl import load_workbook

# ============================== CONFIG ======================================

EXCEL_PATH = os.getenv("EXCEL_PATH", "kaamkibaatein_HR_Contact_Database.xlsx")
SHEET_NAME = "HR Contacts"

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = os.getenv("SMTP_USER", "rishit.tandon.7@gmail.com")   # <-- confirm this is the sending address
SMTP_PASS = os.getenv("SMTP_PASS", "")   # <-- fill in (Gmail App Password, not your login password)

YOUR_NAME = "Rishit Tandon"
YOUR_PHONE = "+91 7394865520"              # leave blank ("") to omit from the email
YOUR_LINKEDIN = "linkedin.com/in/rishit-tandon-928661287"
YOUR_PORTFOLIO = "portfolio.rishit.site"
YOUR_GITHUB = "github.com/RishitTandon7"

RESUME_PATH = os.getenv("RESUME_PATH", "Rishit-Tandon_Resume.pdf")   # <-- put your resume PDF next to this script (or full path)

GEMINI_API_BASE = "https://generativelanguage.googleapis.com/v1beta/models"

GEMINI_MODELS = [
    model.strip()
    for model in os.getenv(
        "GEMINI_MODELS",
        "gemini-3.5-flash,gemini-3-flash,gemini-2.5-flash,gemini-2.5-flash-lite,gemini-2.5-pro",
    ).split(",")
    if model.strip()
]

_GEMINI_MODEL_INDEX = 0

GOOGLE_API_KEYS = [
    os.getenv(f"GOOGLE_API_KEY_{index}")
    for index in range(1, 17)
]
GOOGLE_API_KEYS.extend([
    os.getenv("GOOGLE_API_KEY"),
    os.getenv("GEMINI_API_KEY"),
])
GOOGLE_API_KEYS = [key for key in GOOGLE_API_KEYS if key]

_GOOGLE_API_KEY_INDEX = 0

SENT_LOG_PATH = "sent_log.json"
STATS_PATH = "stats.json"

DAILY_LIMIT = int(os.getenv("DAILY_LIMIT", "40"))    # emails per run — keep this low, raise gradually if no bounces/spam flags
DELAY_SECONDS = int(os.getenv("DELAY_SECONDS", "25"))  # pause between sends within a run

DRY_RUN = os.getenv("DRY_RUN", "true").strip().lower() not in ("false", "0", "no")   # <-- set false only after checking dry-run output

SUBJECT_TEMPLATE = "SRM Final-Year CS Student — AI/ML & Full-Stack — Internship/Full-Time Opportunities at {company}"

# =============================================================================


PROFILE_SUMMARY = f"""
You are helping draft a cold outreach email from a student to an HR contact.
Student profile:
- {YOUR_NAME}, final-year B.Tech Computer Science student at SRM Institute of Science and Technology, graduating 2027
- Focus: AI/ML, Computer Vision, Multimodal AI, IoT, and full-stack development
- Current Research Intern (ML & Image Processing) at SRM IST in collaboration with NIDM, building a satellite-based
  glacial lake outburst flood (GLOF) early-warning pipeline
- Prior Software Development Intern at DGTL Innovations (Python/MySQL REST APIs, frontend work)
- 12+ national/international hackathon wins (1st place at SUTD Smorphi Singapore, CismoHack, and others), plus
  one accepted research publication
- 90+ public GitHub repos, 12,000+ annual GitHub contributions
- Built and shipped: DocMind (agentic Graph RAG platform), an AI agent marketplace (React/FastAPI/Supabase/AWS),
  NeuroVibe (EEG-controlled smart wheelchair), an autonomous AI negotiation system, AI Smart Glasses (VLM-based),
  QML-PLACE (quantum ML VLSI placement engine), and LinkedOut (AI LinkedIn content scheduler)
- Portfolio: {YOUR_PORTFOLIO}, GitHub: {YOUR_GITHUB}, LinkedIn: {YOUR_LINKEDIN}
- Open to both internship and full-time roles
""".strip()


def get_next_google_api_key() -> Optional[str]:
    global _GOOGLE_API_KEY_INDEX
    if not GOOGLE_API_KEYS:
        return None

    api_key = GOOGLE_API_KEYS[_GOOGLE_API_KEY_INDEX % len(GOOGLE_API_KEYS)]
    _GOOGLE_API_KEY_INDEX = (_GOOGLE_API_KEY_INDEX + 1) % len(GOOGLE_API_KEYS)
    return api_key


def get_next_gemini_model() -> Optional[str]:
    global _GEMINI_MODEL_INDEX
    if not GEMINI_MODELS:
        return None

    model = GEMINI_MODELS[_GEMINI_MODEL_INDEX % len(GEMINI_MODELS)]
    _GEMINI_MODEL_INDEX = (_GEMINI_MODEL_INDEX + 1) % len(GEMINI_MODELS)
    return model


def gemini_generate_text(prompt: str) -> str:
    last_error = None
    api_key_attempts = len(GOOGLE_API_KEYS) if GOOGLE_API_KEYS else 0
    model_attempts = len(GEMINI_MODELS) if GEMINI_MODELS else 0

    for _ in range(max(api_key_attempts, 1) * max(model_attempts, 1)):
        api_key = get_next_google_api_key()
        model = get_next_gemini_model()
        if not api_key or not model:
            break

        url = f"{GEMINI_API_BASE}/{model}:generateContent?key={api_key}"
        payload = {
            "contents": [
                {
                    "role": "user",
                    "parts": [{"text": prompt}],
                }
            ],
            "generationConfig": {
                "temperature": 0.7,
                "maxOutputTokens": 80,
            },
        }

        try:
            resp = requests.post(url, json=payload, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            candidates = data.get("candidates", [])
            if candidates:
                content = candidates[0].get("content", {})
                parts = content.get("parts", [])
                text = "".join(part.get("text", "") for part in parts).strip()
                if text:
                    return text
        except Exception as exc:
            last_error = exc
            continue

    if last_error:
        raise last_error
    raise RuntimeError("No Google API keys or Gemini models configured")


def get_personalized_opener(name: str, title: str, company: str) -> str:
    """Ask Gemini for a short, genuine opening line. Falls back to a safe generic line on any failure."""
    first_name = str(name).split()[0] if name else "there"
    prompt = f"""{PROFILE_SUMMARY}

Write ONE short opening line (max 25 words, no greeting like "Hi" or "Dear", no sign-off) for a cold
email to {name}, who is {title} at {company}. It should sound natural and specific to their role/company,
not generic flattery, and should NOT mention hackathons or GitHub stats (save those for later in the email).
Return ONLY the line, no quotes, no explanation.
"""
    try:
        text = gemini_generate_text(prompt).strip()
        text = text.strip('"').strip()
        if text and len(text) < 300:
            return text
    except Exception as e:
        print(f"  [gemini fallback for {first_name}: {e}]")
    return f"I've been following the growth at {company} and wanted to reach out directly."


def build_email_body(name: str, title: str, company: str, opener: str) -> str:
    first_name = str(name).split()[0] if name else "there"
    phone_line = f"\nPhone: {YOUR_PHONE}" if YOUR_PHONE else ""
    return f"""Hi {first_name},

{opener}

I'm {YOUR_NAME}, a final-year Computer Science student at SRM Institute of Science and Technology
(graduating 2027), focused on AI/ML, Computer Vision, and full-stack development. I'm reaching out to ask if
{company} has openings — internship or full-time — where my background would be a fit.

A quick snapshot of what I've built and shipped:
- Currently a Research Intern (ML & Image Processing) at SRM IST x NIDM, building a satellite-based
  early-warning pipeline for glacial lake outburst floods
- Prior SDE internship at DGTL Innovations (Python/MySQL REST APIs)
- 12+ national/international hackathon wins, plus one accepted research publication
- Shipped projects spanning agentic RAG systems (DocMind), a live AI agent marketplace, an EEG-controlled
  assistive wheelchair (NeuroVibe), and a quantum ML VLSI placement engine (QML-PLACE)
- 90+ public GitHub repos, 12,000+ annual contributions

Portfolio: {YOUR_PORTFOLIO}
GitHub: {YOUR_GITHUB}
LinkedIn: {YOUR_LINKEDIN}{phone_line}

I've attached my resume. Would genuinely appreciate the chance to talk, or to be pointed to the
right person/process on your team if this isn't the right fit for you directly.

Thanks for your time,
{YOUR_NAME}
"""


def load_contacts():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    # Some exports tag rows with a "Relevance Tier" column to exclude "Not HR" rows;
    # the plain export doesn't have that column, so fall back to keeping every row.
    has_tier = "Relevance Tier" in col
    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col["Name"]]
        email = row[col["Email"]]
        if not name or not email:
            continue  # section header / incomplete row
        if has_tier and row[col["Relevance Tier"]] not in ("Strong HR - Priority", "HR-Adjacent - Useful"):
            continue
        contacts.append({
            "name": name,
            "title": row[col["Title"]],
            "company": row[col["Company"]],
            "email": email,
        })
    return contacts


def load_sent_log() -> dict:
    if os.path.exists(SENT_LOG_PATH):
        with open(SENT_LOG_PATH) as f:
            data = json.load(f)
        if isinstance(data, list):  # legacy format: plain list of emails
            return {email: {} for email in data}
        return data
    return {}


def append_sent_log(contact: dict):
    sent = load_sent_log()
    sent[contact["email"]] = {
        "name": contact["name"],
        "title": contact["title"],
        "company": contact["company"],
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(SENT_LOG_PATH, "w") as f:
        json.dump(sent, f, indent=2, sort_keys=True)


def write_stats(total_contacts: int, sent: dict, run_result: str):
    today = datetime.now(timezone.utc).date().isoformat()
    sent_today = sum(1 for v in sent.values() if isinstance(v, dict) and str(v.get("sent_at", "")).startswith(today))
    stats = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_contacts": total_contacts,
        "total_sent": len(sent),
        "remaining": max(total_contacts - len(sent), 0),
        "sent_today": sent_today,
        "daily_limit": DAILY_LIMIT,
        "dry_run": DRY_RUN,
        "gemini_models": GEMINI_MODELS,
        "google_api_key_count": len(GOOGLE_API_KEYS),
        "last_run_result": run_result,
    }
    with open(STATS_PATH, "w") as f:
        json.dump(stats, f, indent=2)


def send_email(to_email: str, subject: str, body: str):
    msg = MIMEMultipart()
    msg["From"] = SMTP_USER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    if RESUME_PATH and os.path.exists(RESUME_PATH):
        with open(RESUME_PATH, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(RESUME_PATH))
        part["Content-Disposition"] = f'attachment; filename="{os.path.basename(RESUME_PATH)}"'
        msg.attach(part)

    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls(context=context)
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, to_email, msg.as_string())


def main():
    contacts = load_contacts()
    sent = load_sent_log()
    pending = [c for c in contacts if c["email"] not in sent]
    print(f"Total useful contacts: {len(contacts)} | Already sent: {len(sent)} | Remaining: {len(pending)}")
    print(f"Model rotation: {len(GEMINI_MODELS)} Gemini models x {len(GOOGLE_API_KEYS)} API keys")

    if not pending:
        print("Nothing left to send. All useful contacts have been emailed.")
        write_stats(len(contacts), sent, run_result="nothing_pending")
        return

    batch = pending[:3] if DRY_RUN else pending[:DAILY_LIMIT]
    sent_count = 0

    for i, c in enumerate(batch, 1):
        opener = get_personalized_opener(c["name"], c["title"], c["company"])
        body = build_email_body(c["name"], c["title"], c["company"], opener)
        subject = SUBJECT_TEMPLATE.format(company=c["company"])

        print(f"\n--- [{i}/{len(batch)}] {c['name']} | {c['title']} | {c['company']} | {c['email']} ---")
        print(f"Subject: {subject}")
        print(body)

        if DRY_RUN:
            continue

        try:
            send_email(c["email"], subject, body)
            append_sent_log(c)
            sent = load_sent_log()
            sent_count += 1
            print(f"  -> sent to {c['email']}")
        except Exception as e:
            print(f"  -> FAILED for {c['email']}: {e}")

        if i < len(batch):
            time.sleep(DELAY_SECONDS)

    if DRY_RUN:
        print("\n\nDRY RUN complete — nothing was sent. Review the 3 emails above, then set DRY_RUN = false to actually send.")
        write_stats(len(contacts), sent, run_result="dry_run")
    else:
        print(f"\n\nRun complete. Sent {sent_count} emails this run. Rerun the script tomorrow for the next batch.")
        write_stats(len(contacts), sent, run_result=f"sent_{sent_count}")


if __name__ == "__main__":
    main()
